import base64
import io
import json
import os
import re
import socket
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import openpyxl
import requests
from bs4 import BeautifulSoup
from flask import (Response, abort, Flask, jsonify, render_template_string,
                   request, send_file)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# disable insecure warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------- Config ----------
app = Flask(__name__)
executor = ThreadPoolExecutor(max_workers=10)

VERSION_RE = re.compile(r"\b\d+(?:\.\d+){0,}\b")
PROD_SLASH_VER_RE = re.compile(r"([A-Za-z0-9\-]+)/(\d+[A-Za-z0-9\.\-_]*)")
LEADING_V_RE = re.compile(r"\bv\d+(?:\.\d+)*\b", re.IGNORECASE)

DEFAULT_HEADERS = [
    "Server", "X-Powered-By", "X-AspNet-Version", "X-AspNetMvc-Version",
    "X-Powered-By-Plesk", "X-Powered-By-PHP", "X-PHP-Version", "X-Generator",
    "X-Drupal-Cache", "X-Drupal-Cache-Context", "X-Backend-Server", "X-Backend",
    "X-App-Server", "X-Service", "X-Instance-Id", "X-Instance", "X-Served-By",
    "X-Cache", "X-Cache-Lookup", "Via", "Forwarded", "X-Forwarded-For",
    "Server-Timing", "X-Request-ID", "X-Runtime", "X-Varnish", "X-Amz-Cf-Id",
]

JOBS = {}

# ---------- Load logo as base64 data URI (if present) ----------
def load_logo_data_uri(path="static/logo.png"):
    try:
        with open(path, "rb") as fh:
            data = fh.read()
            b64 = base64.b64encode(data).decode("ascii")
            # Try to infer mime type from extension
            ext = os.path.splitext(path)[1].lower()
            mime = "image/png"
            if ext in (".jpg", ".jpeg"):
                mime = "image/jpeg"
            elif ext == ".svg":
                mime = "image/svg+xml"
            return f"data:{mime};base64,{b64}"
    except Exception:
        return ""

LOGO_DATA_URI = load_logo_data_uri()

# ---------- Helper functions ----------
def has_version_info(value: str):
    if not value:
        return False, "", None, None
    m = PROD_SLASH_VER_RE.search(value)
    if m:
        return True, f"{m.group(1)}/{m.group(2)}", m.group(1), m.group(2)
    m2 = LEADING_V_RE.search(value)
    if m2:
        return True, m2.group(0), None, m2.group(0)
    m3 = VERSION_RE.search(value)
    if m3:
        return True, m3.group(0), None, m3.group(0)
    return False, "", None, None

def severity_for(header_name: str, has_version: bool):
    h = header_name.lower()
    if has_version:
        if h == "server":
            return "High"
        if "x-powered" in h or h.startswith("x-aspnet") or "php" in h:
            return "Medium"
    return "Low"

def cve_search_hint(product: str, version: str):
    if product and version:
        return f"{product} {version} vulnerabilities"
    if version:
        return f"vulnerabilities {version}"
    return ""

def lookup_cves(product: str, version: str, limit: int = 5):
    if not product and not version:
        return []
    token = (f"{product}/{version}" if product and version else (product or version)).lower().replace(" ", "/")
    url = f"https://cve.circl.lu/api/search/{token}"
    try:
        r = requests.get(url, headers={"User-Agent":"exposure-scan/1.0"}, timeout=6)
        if r.status_code == 200:
            items = r.json()
            out = []
            for node in items[:limit]:
                out.append({"id": node.get("id"), "summary": node.get("summary")})
            return out
    except Exception:
        pass
    return []

def sanitize_value_for_xlsx(val):
    if val is None:
        return ""
    try:
        text = BeautifulSoup(val, "html.parser").get_text(separator=" ", strip=True)
    except Exception:
        text = str(val)
    text = re.sub(r'\s+', ' ', text).strip()
    if len(text) > 20000:
        text = text[:20000] + " ... (truncated)"
    return text

# ---------- Scanner ----------
def scan_single_target(url, headers_to_check=None, verify_ssl=True, timeout=10, opts=None, job_id=None):
    if opts is None:
        opts = {}
    result = {"url": url, "timestamp": datetime.utcnow().isoformat()+"Z", "vulnerable": False, "findings": [], "notes": []}

    if headers_to_check is None:
        headers_to_check = DEFAULT_HEADERS.copy()
    if not opts.get("include_generator", False):
        headers_to_check = [h for h in headers_to_check if h.lower() != "x-generator"]

    # DNS quick-check
    try:
        domain = url.replace("https://","").replace("http://","").split("/")[0]
        socket.gethostbyname(domain)
    except Exception as e:
        result["notes"].append(f"DNS failed: {e}")
        return result

    # GET main page
    try:
        resp = requests.get(url, timeout=timeout, verify=verify_ssl)
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url, timeout=timeout, verify=False)
            result["notes"].append("SSL verify failed; retried insecure")
        except Exception as e:
            result["notes"].append(f"Request failed: {e}")
            return result
    except Exception as e:
        result["notes"].append(f"Request failed: {e}")
        return result

    # Inspect headers (only if version evidence present)
    for h in headers_to_check:
        val = resp.headers.get(h)
        if not val:
            continue
        is_ver, evidence, prod, ver = has_version_info(val)
        if not is_ver:
            continue
        sev = severity_for(h, is_ver)
        fnd = {
            "header": h,
            "value": val,
            "version_present": is_ver,
            "evidence": evidence,
            "product": prod,
            "version": ver,
            "severity": sev,
            "cve_search_hint": cve_search_hint(prod, ver),
            "source": "header",
            "cves": []
        }
        if prod or ver:
            try:
                cves = lookup_cves(prod, ver)
                if cves:
                    fnd["cves"] = cves
                    fnd["severity"] = "High"
            except Exception:
                pass
        result["findings"].append(fnd)
        result["vulnerable"] = True

    # Active endpoints
    if opts.get("active", False):
        endpoints = ["/server-status","/status","/phpinfo.php","/robots.txt","/.git/config","/wp-login.php","/admin","/.env"]
        for ep in endpoints:
            if job_id and JOBS.get(job_id, {}).get("cancelled"):
                result["notes"].append("scan cancelled")
                return result
            try:
                hresp = requests.head(url.rstrip("/") + ep, timeout=6, verify=verify_ssl, allow_redirects=True)
                if hresp.status_code and hresp.status_code < 400:
                    result["findings"].append({
                        "header":"active:head",
                        "value": f"{ep} -> {hresp.status_code}",
                        "version_present": False,
                        "evidence": None,
                        "product": None,
                        "version": None,
                        "severity": "Low",
                        "cve_search_hint": "active endpoint info",
                        "source": "active"
                    })
            except Exception:
                pass

    # Deep HTML checks
    if opts.get("deep", True):
        try:
            ctype = resp.headers.get("Content-Type","")
            body = resp.text or ""
            if ("html" in ctype.lower()) or body:
                soup = BeautifulSoup(body, "html.parser")
                meta = soup.find("meta", attrs={"name": re.compile(r"generator", re.I)})
                if meta and meta.get("content"):
                    val = meta["content"]
                    is_ver, evidence, prod, ver = has_version_info(val)
                    mf = {
                        "header":"meta:generator",
                        "value": val,
                        "version_present": is_ver,
                        "evidence": evidence,
                        "product": prod,
                        "version": ver,
                        "severity": "Medium" if is_ver else "Low",
                        "cve_search_hint": cve_search_hint(prod, ver),
                        "source":"html-meta",
                        "cves": []
                    }
                    if prod or ver:
                        try:
                            mf["cves"] = lookup_cves(prod, ver)
                            if mf["cves"]:
                                mf["severity"] = "High"
                        except Exception:
                            pass
                    result["findings"].append(mf)
                    if is_ver:
                        result["vulnerable"] = True

                comments = re.findall(r"<!--([\s\S]*?)-->", body)
                interest_keywords = ["admin","login","wp-admin","backup",".env","api_key","api-key","secret","password","passwd","sql error","exception","traceback","ora-","mysql","jdbc"]
                for c in comments:
                    lowered = c.lower()
                    is_interesting = any(k in lowered for k in interest_keywords)
                    m = PROD_SLASH_VER_RE.search(c) or LEADING_V_RE.search(c) or VERSION_RE.search(c)
                    evidence = None
                    if m:
                        evidence = m.group(0)
                        if re.fullmatch(r"\d", evidence):
                            evidence = None
                        if evidence and not ('.' in evidence or evidence.startswith('v') or '/' in evidence):
                            evidence = None
                    if not (is_interesting or evidence):
                        continue
                    sev = "Medium" if is_interesting else "Low"
                    if is_interesting and any(k in lowered for k in ['api_key','.env','passwd','password','secret']):
                        sev = "High"
                    comment_f = {
                        "header":"html:comment",
                        "value": c.strip()[:1200],
                        "version_present": bool(evidence),
                        "evidence": evidence,
                        "product": None,
                        "version": evidence,
                        "severity": sev,
                        "cve_search_hint": cve_search_hint(None, evidence) if evidence else "",
                        "source": "html-comment"
                    }
                    result["findings"].append(comment_f)
                    result["vulnerable"] = True
                    break
        except Exception as e:
            result["notes"].append(f"HTML parse error: {e}")

        # robots.txt
        try:
            robots_url = "/".join(url.split("/")[:3]) + "/robots.txt"
            r2 = requests.get(robots_url, timeout=6, verify=verify_ssl)
            if r2.status_code == 200 and r2.text:
                lines = [l.strip() for l in r2.text.splitlines() if l.strip()]
                sensitive = [l for l in lines if re.search(r"(admin|config|wp-admin|.env|backup|backup\-db|private|secret)", l, re.I)]
                if sensitive:
                    result["findings"].append({
                        "header":"robots.txt",
                        "value": "\n".join(sensitive[:8]),
                        "version_present": False,
                        "severity": "Medium",
                        "cve_search_hint": "robots.txt disallow sensitive paths",
                        "source": "robots"
                    })
                    result["vulnerable"] = True
        except Exception:
            pass

        # sitemap.xml
        try:
            sitemap_url = "/".join(url.split("/")[:3]) + "/sitemap.xml"
            r3 = requests.get(sitemap_url, timeout=6, verify=verify_ssl)
            if r3.status_code == 200 and r3.text:
                m = re.search(r"(http[s]?://[\w\-.]*internal|http[s]?://[\w\-.]*staging|http[s]?://[\w\-.]*dev)", r3.text, re.I)
                if m:
                    result["findings"].append({
                        "header":"sitemap.xml",
                        "value": m.group(0),
                        "version_present": False,
                        "severity": "Medium",
                        "cve_search_hint": "sitemap internal host leak",
                        "source": "sitemap"
                    })
                    result["vulnerable"] = True
        except Exception:
            pass

    return result

# ---------- Job runner ----------
def _run_scan_job(job_id, targets, opts):
    JOBS[job_id] = JOBS.get(job_id, {})
    JOBS[job_id]['status'] = 'running'
    JOBS[job_id]['started'] = datetime.utcnow().isoformat() + "Z"
    JOBS[job_id]['scanned'] = 0
    total = len([t for t in targets if t and t.strip()])
    JOBS[job_id]['total'] = total
    JOBS[job_id]['results'] = []
    JOBS[job_id]['cancelled'] = False

    for t in targets:
        if JOBS[job_id].get('cancelled'):
            JOBS[job_id]['status'] = 'cancelled'
            JOBS[job_id]['finished'] = datetime.utcnow().isoformat() + "Z"
            return
        u = t.strip()
        if not u:
            JOBS[job_id]['scanned'] = len(JOBS[job_id]['results'])
            continue
        if not u.startswith("http://") and not u.startswith("https://"):
            u = "https://" + u
        r = scan_single_target(u, headers_to_check=opts.get('headers'), verify_ssl=not opts.get('no_verify', False), timeout=opts.get('timeout',10), opts=opts, job_id=job_id)
        JOBS[job_id]['results'].append(r)
        JOBS[job_id]['scanned'] = len(JOBS[job_id]['results'])

    JOBS[job_id]['status'] = 'finished'
    JOBS[job_id]['finished'] = datetime.utcnow().isoformat() + "Z"

# ---------- UI / Templates ----------
INDEX_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title style="margin:4rm">INFOSIGHT</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.5/font/bootstrap-icons.css" rel="stylesheet">
<style>
:root{--bg:#f6f7fb;--card:#fff;--muted:#6c757d;--text:#111}
body.dark{--bg:#0b0d0f;--card:#0f1416;--muted:#9aa0a6;--text:#e6edf3;background:var(--bg);color:var(--text)}
body{background:var(--bg);color:var(--text)}
.container{max-width:1100px}
.dropzone{border:2px dashed rgba(108,117,125,0.15);padding:14px;border-radius:8px;background:transparent}
.dropzone.dragover{border-color:#6c757d;background:rgba(108,117,125,0.03)}
.findings-table{overflow:auto;max-width:100%}
pre{white-space:pre-wrap;word-break:break-word;background:transparent;border:0;margin:0;padding:0}
.badge-high{background:#dc3545;color:white}
.badge-med{background:#ffc107;color:#111}
.badge-low{background:#6c757d;color:white}
.btn-theme{position:fixed;right:18px;top:18px;z-index:1100}
.small-note{color:var(--muted)}
.header-logo{height:48px;margin-right:12px}
@media (max-width:576px){ .header-logo{height:36px;margin-right:8px} }
img { filter: none !important; }
</style>
</head>
<body>
<button id="themeBtn" class="btn btn-sm btn-outline-secondary btn-theme"><i class="bi bi-moon-stars"></i></button>
<div class="container py-4">
  <div class="d-flex align-items-center mb-3 flex-wrap">
    {% if logo_data_uri %}
      <img src="{{ logo_data_uri }}" class="header-logo" alt="Exposure Scan logo">
    {% endif %}
    <h1 class="mb-10" style="padding-left:350px; font-weight: bold; font-style: italic; font-size: 5em;color:red">INFOSIGHT</h1>     
  </div>
  <p class="small-note">Upload a .txt (1 target per line) or paste targets. Deep scan is <strong>ON by default</strong>.</p>

  <form id="scanForm" class="mb-3">
    <div class="row g-2">
      <div class="col-md-6">
        <label class="form-label">Job Name (optional)</label>
        <input id="jobName" class="form-control" placeholder="friendly label (optional)" />
      </div>
      <div class="col-md-6">
        <label class="form-label">Upload .txt (drag & drop)</label>
        <div id="dropzone" class="dropzone">
          <input id="fileInput" type="file" accept=".txt" class="form-control" />
          <div id="filePreview" class="mt-2 small-note">No file selected</div>
        </div>
      </div>
    </div>

    <div class="mb-3 mt-3">
      <label class="form-label">Or paste targets (one per line)</label>
      <textarea id="targets" rows="5" class="form-control" placeholder="https://example.com"></textarea>
    </div>

    <div class="d-flex gap-3 mb-3">
      <div class="form-check">
        <input id="deep" class="form-check-input" type="checkbox" checked>
        <label class="form-check-label" for="deep">Deep scan (enabled by default)</label>
      </div>
      <div class="form-check">
        <input id="active" class="form-check-input" type="checkbox">
        <label class="form-check-label" for="active">Active fingerprinting (HEAD)</label>
      </div>
      <div class="form-check">
        <input id="noverify" class="form-check-input" type="checkbox">
        <label class="form-check-label" for="noverify">Skip SSL verify</label>
      </div>
    </div>

    <div>
      <button id="startBtn" class="btn btn-primary"><i class="bi bi-play-fill"></i> Start Scan</button>
      <button id="stopBtn" type="button" class="btn btn-danger" style="display:none; margin-left:8px;"><i class="bi bi-stop-fill"></i> Stop Scan</button>
    </div>
  </form>

  <div id="jobPanel" class="mb-3"></div>

  <div class="mb-3">
    <div class="progress" style="height:22px; display:none" id="progWrap">
      <div id="progBar" class="progress-bar" role="progressbar" style="width:0%">0%</div>
    </div>
    <div id="progText" class="small-note mt-1"></div>
  </div>

  <hr>
  <h5>Recent jobs</h5>
  <div id="recent" class="mb-3"></div>

  <h5>Vulnerable-only preview (last job)</h5>
  <div id="vulnPreview"></div>
</div>

<script>
const dropzone = document.getElementById('dropzone');
const fileInput = document.getElementById('fileInput');
const filePreview = document.getElementById('filePreview');
const startBtn = document.getElementById('startBtn');
const stopBtn = document.getElementById('stopBtn');
let currentJob = null;

['dragenter','dragover'].forEach(e=>dropzone.addEventListener(e, ev=>{ev.preventDefault(); dropzone.classList.add('dragover');}));
['dragleave','drop'].forEach(e=>dropzone.addEventListener(e, ev=>{ev.preventDefault(); dropzone.classList.remove('dragover');}));
dropzone.addEventListener('drop', async ev=>{ const f = ev.dataTransfer.files[0]; if(f){ fileInput.files = ev.dataTransfer.files; await showPreview(f);} });
fileInput.addEventListener('change', async ()=>{ if(fileInput.files.length>0) await showPreview(fileInput.files[0]); });

async function showPreview(f){
  try {
    const text = await f.text();
    const lines = text.split(/\\r?\\n/).slice(0,10);
    filePreview.innerHTML = '<strong>Preview:</strong><pre>'+lines.join('\\n')+'</pre>';
  } catch(e){
    filePreview.innerHTML = '<small class="text-danger">Cannot preview file</small>';
  }
}

document.getElementById('scanForm').addEventListener('submit', startScan);
stopBtn.addEventListener('click', cancelJob);

async function startScan(e){
  e.preventDefault();
  const jobName = document.getElementById('jobName').value.trim();
  const opts = {
    deep: document.getElementById('deep').checked,
    active: document.getElementById('active').checked,
    no_verify: document.getElementById('noverify').checked,
    include_generator: false,
    job_name: jobName || ""
  };
  const form = new FormData();
  form.append('opts', JSON.stringify(opts));
  if(fileInput.files.length>0){
    form.append('file', fileInput.files[0]);
  } else {
    const t = document.getElementById('targets').value.trim();
    if(!t){ alert('Add a file or paste targets'); return; }
    form.append('targets', t);
  }

  document.getElementById('jobPanel').innerHTML = '';
  document.getElementById('vulnPreview').innerHTML = '';
  document.getElementById('progWrap').style.display = 'block';
  document.getElementById('progBar').style.width = '0%';
  document.getElementById('progBar').textContent = '0%';
  document.getElementById('progText').textContent = '';

  const r = await fetch('/api/scan', { method:'POST', body: form });
  const j = await r.json();
  currentJob = j.job_id;
  stopBtn.style.display = 'inline-block';
  pollStatus(currentJob);
}

async function cancelJob(){
  if(!currentJob) return;
  await fetch('/api/cancel/'+currentJob, { method:'POST' });
  document.getElementById('jobPanel').innerHTML += `<div class="mt-2 small-note">Cancel requested</div>`;
}

async function pollStatus(job_id){
  const r = await fetch('/api/status/'+job_id);
  const j = await r.json();
  const short = j.job_name && j.job_name.length ? j.job_name : job_id.slice(0,8);
  let panel = `<div><strong>Job ${short}</strong> — status: <strong>${j.status}</strong></div>`;
  if(j.started) panel += `<div class="small-note">Started: ${new Date(j.started).toLocaleString()}</div>`;
  if(j.finished) panel += `<div class="small-note">Finished: ${new Date(j.finished).toLocaleString()}</div>`;
  document.getElementById('jobPanel').innerHTML = panel;

  const scanned = j.scanned || 0, total = j.total || 0;
  const pct = total>0 ? Math.round((scanned/total)*100) : 0;
  document.getElementById('progBar').style.width = pct + '%';
  document.getElementById('progBar').textContent = pct + '%';
  document.getElementById('progText').textContent = `Scanned ${scanned} of ${total}`;

  if(j.status === 'running' || j.status === 'queued'){
    setTimeout(()=>pollStatus(job_id), 800);
  } else {
    stopBtn.style.display = 'none';
    document.getElementById('jobPanel').innerHTML += `<div class="mt-2"><a class="btn btn-sm btn-success" href="/report/${job_id}?format=html" target="_blank">Open HTML</a> <a class="btn btn-sm btn-secondary" href="/report/${job_id}?format=xlsx">XLSX</a> <a class="btn btn-sm btn-secondary" href="/report/${job_id}?format=zip">ZIP</a></div>`;
    fetchVulnPreview(job_id);
    refreshRecent();
  }
}

async function refreshRecent(){
  const r = await fetch('/api/recent');
  const j = await r.json();
  const el = document.getElementById('recent');
  if(!j.length){ el.innerHTML = '<em>No recent jobs</em>'; return; }
  let html = '<ul class="list-group">';
  j.forEach(job => {
    const short = job.job_name && job.job_name.length ? job.job_name : job.job_id.slice(0,8);
    html += `<li class="list-group-item d-flex justify-content-between align-items-center">${short} <span class="badge ${job.status=='finished'?'bg-success':'bg-secondary'}">${job.status}</span></li>`;
  });
  html += '</ul>';
  el.innerHTML = html;
}

async function fetchVulnPreview(job_id){
  const r = await fetch('/api/results/'+job_id);
  if(r.status !== 200){ document.getElementById('vulnPreview').innerHTML = '<em>No vulnerable targets or results not ready</em>'; return; }
  const j = await r.json();
  if(!j.length){ document.getElementById('vulnPreview').innerHTML = '<em>No vulnerable targets</em>'; return; }
  let out = '<ol>';
  j.forEach(v => out += `<li><strong>${v.url}</strong> — ${v.findings.length} findings</li>`);
  out += '</ol>';
  document.getElementById('vulnPreview').innerHTML = out;
}

// theme persistence
const themeBtn = document.getElementById('themeBtn');
(function initTheme(){
  try {
    const t = localStorage.getItem('theme');
    if(t === 'dark') document.body.classList.add('dark');
    themeBtn.innerHTML = document.body.classList.contains('dark') ? '<i class="bi bi-sun-fill"></i>' : '<i class="bi bi-moon-stars"></i>';
  } catch(e){}
})();
themeBtn.addEventListener('click', ()=>{
  document.body.classList.toggle('dark');
  themeBtn.innerHTML = document.body.classList.contains('dark') ? '<i class="bi bi-sun-fill"></i>' : '<i class="bi bi-moon-stars"></i>';
  try { localStorage.setItem('theme', document.body.classList.contains('dark') ? 'dark' : 'light'); } catch(e){}
});

refreshRecent();
</script>
</body>
</html>
"""

# ---------- Reports / XLSX ----------
REPORT_TEMPLATE = """
<!doctype html><html><head><meta charset="utf-8"><title>Exposure Scan Report - {{ job.job_id }}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head><body class="p-4"><div class="container">
  <div class="d-flex align-items-center mb-3">
    {% if logo_data_uri %}
      <img src="{{ logo_data_uri }}" style="height:42px; margin-right:10px;" alt="Exposure Scan logo">
    {% endif %}
    <h1 class="mb-0">Exposure Scan Report</h1>
  </div>
  <p>Job: <strong>{{ job.job_name if job.job_name else job.job_id[:8] }}</strong> — status: <strong>{{ job.status }}</strong></p>
  <p>Started: {{ job.started }} | Finished: {{ job.finished }}</p>
  {% for r in job.results %}
    <div class="card mb-3"><div class="card-body">
      <h5 class="card-title">{{ r.url }} <small class="text-muted">Vulnerable: {{ 'Yes' if r.vulnerable else 'No' }}</small></h5>
      {% if r.notes %}<p><strong>Notes:</strong> {{ r.notes|join(', ') }}</p>{% endif %}
      {% if r.findings %}
        <div class="table-responsive"><table class="table table-sm"><thead><tr><th>Header</th><th>Value</th><th>Evidence</th><th>Severity</th><th>CVEs / Hints</th></tr></thead><tbody>
        {% for f in r.findings %}
          <tr>
            <td style="min-width:120px">{{ f.header }}</td>
            <td style="min-width:320px"><pre>{{ f.value }}</pre></td>
            <td>{{ f.evidence or '' }}</td>
            <td>{{ f.severity }}</td>
            <td>
              {% if f.cves %}
                <ul class="small mb-0">{% for c in f.cves %}<li><a target="_blank" href="https://cve.circl.lu/cve/{{ c.id }}">{{ c.id }}</a> - {{ c.summary[:140] }}{% if c.summary|length > 140 %}...{% endif %}</li>{% endfor %}</ul>
              {% elif f.cve_search_hint %}
                <a target="_blank" href="https://www.google.com/search?q={{ f.cve_search_hint|urlencode }}">Search</a>
              {% endif %}
            </td>
          </tr>
        {% endfor %}
        </tbody></table></div>
      {% else %}
        <p class="text-muted">No findings.</p>
      {% endif %}
    </div></div>
  {% endfor %}
  <hr>
  <footer class="text-muted small">Powered by <strong>Exposure Scan</strong></footer>
</div></body></html>
"""

def make_xlsx_for_job(job):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = ["url_index","url","vulnerable","header","value","evidence","product","version","severity","cve_search_hint"]
    ws.append(headers)
    for col_idx in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    url_index = {}
    idx = 1
    for r in job.get("results", []):
        u = r["url"]
        if u not in url_index:
            url_index[u] = idx
            idx += 1

    fill_high = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_med = PatternFill(start_color="FFF7CC", end_color="FFF7CC", fill_type="solid")
    fill_low = PatternFill(start_color="E9EEF2", end_color="E9EEF2", fill_type="solid")
    row = 2
    for r in job.get("results", []):
        uidx = url_index[r["url"]]
        for f in r.get("findings", []):
            val = sanitize_value_for_xlsx(f.get("value"))
            evidence = sanitize_value_for_xlsx(f.get("evidence"))
            product = f.get("product") or ""
            version = f.get("version") or ""
            severity = f.get("severity") or "Low"
            hint = f.get("cve_search_hint") or ""
            ws.cell(row=row, column=1, value=uidx)
            ws.cell(row=row, column=2, value=r["url"])
            ws.cell(row=row, column=3, value="Yes" if r.get("vulnerable") else "No")
            ws.cell(row=row, column=4, value=f.get("header"))
            ws.cell(row=row, column=5, value=val)
            ws.cell(row=row, column=6, value=evidence)
            ws.cell(row=row, column=7, value=product)
            ws.cell(row=row, column=8, value=version)
            ws.cell(row=row, column=9, value=severity)
            ws.cell(row=row, column=10, value=hint)
            for c in (4,5,6,10):
                ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            sev_cell = ws.cell(row=row, column=9)
            sev_lower = severity.lower()
            if sev_lower == "high":
                sev_cell.fill = fill_high
            elif sev_lower == "medium":
                sev_cell.fill = fill_med
            else:
                sev_cell.fill = fill_low
            row += 1

    max_cols = min(ws.max_column, 12)
    for col in range(1, max_cols+1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = str(cell.value or "")
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        width = min(max(8, int(max_len * 0.12) + 6), 80)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    try:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ---------- API / Endpoints ----------
@app.route("/")
def index():
    return render_template_string(INDEX_HTML, logo_data_uri=LOGO_DATA_URI)

@app.route("/api/scan", methods=["POST"])
def api_scan():
    opts = {}
    targets = []
    if "file" in request.files:
        f = request.files["file"]
        try:
            content = f.stream.read().decode("utf-8", errors="ignore")
            targets = [line.strip() for line in content.splitlines() if line.strip()]
        except Exception as e:
            return jsonify({"error":"failed to read file","details": str(e)}), 400
        opts_raw = request.form.get("opts")
        try:
            opts = json.loads(opts_raw) if opts_raw else {}
        except Exception:
            opts = {}
    else:
        if request.form.get("targets"):
            raw = request.form.get("targets")
            targets = [line.strip() for line in raw.splitlines() if line.strip()]
            opts_raw = request.form.get("opts")
            try:
                opts = json.loads(opts_raw) if opts_raw else {}
            except Exception:
                opts = {}
        else:
            payload = request.get_json(force=True, silent=True) or {}
            targets = payload.get("targets", [])
            opts = payload.get("opts", {})

    if not targets:
        return jsonify({"error":"no targets provided"}), 400
    if len(targets) > 5000:
        return jsonify({"error":"too many targets (max 5000)"}), 400

    if "deep" not in opts:
        opts["deep"] = True

    job_id = str(uuid.uuid4())
    job_name = opts.get("job_name") or ""
    JOBS[job_id] = {"status":"queued","job_id":job_id,"job_name":job_name,"opts":opts,"results":None,"scanned":0,"total":0,"cancelled":False}
    opts["headers"] = opts.get("headers") or DEFAULT_HEADERS

    executor.submit(_run_scan_job, job_id, targets, opts)
    return jsonify({"job_id": job_id, "status":"queued"})

@app.route("/api/status/<job_id>")
def api_status(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error":"unknown job"}), 404
    return jsonify({
        "job_id": job_id,
        "job_name": job.get("job_name",""),
        "status": job.get("status"),
        "scanned": job.get("scanned",0),
        "total": job.get("total",0),
        "started": job.get("started"),
        "finished": job.get("finished")
    })

@app.route("/api/cancel/<job_id>", methods=["POST"])
def api_cancel(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error":"unknown job"}), 404
    if job.get("status") not in ("running","queued"):
        return jsonify({"status":"not-cancellable","status_msg":"Job already finished or not running"}), 400
    job["cancelled"] = True
    return jsonify({"status":"cancelling","status_msg":"Cancellation requested; worker will stop shortly"})

@app.route("/api/results/<job_id>")
def api_results(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error":"unknown job"}), 404
    if job.get("status") != "finished":
        return jsonify({"error":"job not finished","status":job.get("status")}), 202
    vuln_only = []
    for r in job.get("results", []):
        if r.get("vulnerable"):
            vuln_only.append({"url": r["url"], "findings": r["findings"]})
    return jsonify(vuln_only)

@app.route("/api/recent")
def api_recent():
    out = sorted(list(JOBS.values()), key=lambda x: x.get('started') or "", reverse=True)[:10]
    summary = [{"job_id": j["job_id"], "status": j["status"], "job_name": j.get("job_name","")} for j in out]
    return jsonify(summary)

@app.route("/report/<job_id>")
def report(job_id):
    fmt = request.args.get("format","html")
    job = JOBS.get(job_id)
    if not job:
        abort(404)
    if job.get("status") != "finished":
        return f"Job {job_id} not finished. Current status: {job.get('status')}", 202
    if fmt == "html":
        return render_template_string(REPORT_TEMPLATE, job=job, logo_data_uri=LOGO_DATA_URI)
    if fmt == "json":
        buf = io.BytesIO()
        buf.write(json.dumps(job.get("results", []), indent=2).encode("utf-8"))
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"exposurescan_{job_id}.json", mimetype="application/json")
    if fmt == "xlsx":
        xlsx_io = make_xlsx_for_job(job)
        return send_file(xlsx_io, as_attachment=True, download_name=f"exposurescan_{job_id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if fmt == "zip":
        html_bytes = render_template_string(REPORT_TEMPLATE, job=job, logo_data_uri=LOGO_DATA_URI).encode("utf-8")
        xlsx_io = make_xlsx_for_job(job)
        zipbuf = io.BytesIO()
        with zipfile.ZipFile(zipbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"exposurescan_{job_id}.html", html_bytes)
            zf.writestr(f"exposurescan_{job_id}.xlsx", xlsx_io.getvalue())
        zipbuf.seek(0)
        return send_file(zipbuf, as_attachment=True, download_name=f"exposurescan_{job_id}.zip", mimetype="application/zip")
    return render_template_string(REPORT_TEMPLATE, job=job, logo_data_uri=LOGO_DATA_URI)

# ---------- Startup ----------
if __name__ == "__main__":
    # demo job for UI when first loaded
    demo = str(uuid.uuid4())
    JOBS[demo] = {"status":"finished","job_id":demo,"job_name":"demo","started":datetime.utcnow().isoformat()+"Z","finished":datetime.utcnow().isoformat()+"Z","results":[],"scanned":0,"total":0,"opts":{},"cancelled":False}
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)